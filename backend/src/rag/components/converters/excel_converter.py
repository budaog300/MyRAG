import time
import logging
import asyncio
from pathlib import Path
from typing import Set, Optional
import io
import zipfile
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from docling.datamodel.base_models import InputFormat, DocumentStream
from docling.document_converter import DocumentConverter

from src.rag.components.converters import BaseDocumentConverter
from src.core.exceptions.converter_exceptions import (
    CorruptedExcelFileError,
    DocumentConversionError,
    DocumentFileNotFoundError,
    EmptyDocumentError,
    PipelineInitializationError,
    UnsupportedFileFormatError,
)

logger = logging.getLogger(__name__)


class ExcelConverter(BaseDocumentConverter):
    """Конвертер таблиц Excel (.xlsx) с порезанной поэлементной обработкой листов через Docling."""

    SUPPORTED_EXTENSIONS: Set[str] = {".xlsx"}
    DELIMITER: str = "\n\n---\n\n"

    def __init__(self):
        try:
            self._docling_converter = DocumentConverter()
        except Exception as exc:
            logger.error("Не удалось инициализировать Docling в ExcelConverter: %s", exc)
            raise PipelineInitializationError(reason=str(exc)) from exc

    def _process_single_sheet_with_docling(
        self,
        wb_original: openpyxl.Workbook,
        sheet_name: str,
        file_path: Path,
    ) -> Optional[str]:
        """
        Изолирует один лист во временный файловый поток (BytesIO),
        прогоняет его через Docling и добавляет контекстный заголовок.
        """
        ws = wb_original[sheet_name]

        if ws.sheet_state != "visible":
            logger.debug(f"Пропущен скрытый лист: {sheet_name}")
            return None

        if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None:
            logger.debug(f"Пропущен пустой лист: {sheet_name}")
            return None

        single_sheet_wb = openpyxl.Workbook()
        single_sheet_wb.remove(single_sheet_wb.active)

        new_ws = single_sheet_wb.create_sheet(title=sheet_name)
        for row in ws.iter_rows(values_only=False):
            new_ws.append([cell.value for cell in row])

        buffer = io.BytesIO()
        single_sheet_wb.save(buffer)
        buffer.seek(0)
        file_name = file_path.name

        try:
            doc_stream = DocumentStream(name=f"{sheet_name}.xlsx", stream=buffer)
            docling_result = self._docling_converter.convert(doc_stream)
            raw_markdown = docling_result.document.export_to_markdown()

            context_header = (
                f"# Файл: {file_name}\n"
                f"## Лист: {sheet_name}\n\n"
                f"> Context: Данные таблицы относятся к документу '{file_name}', раздел '{sheet_name}'.\n\n"
            )

            return context_header + raw_markdown

        except Exception as exc:
            logger.warning(f"Ошибка обработки листа '{sheet_name}' в файле '{file_name}' через Docling: {exc}")
            return None

        finally:
            buffer.close()
            single_sheet_wb.close()

    def _preprocess_excel(self, file_path: Path) -> str:
        """Синхронный обход всех листов Excel и сборка в единую строку."""
        try:
            wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        except (InvalidFileException, zipfile.BadZipFile, KeyError) as exc:
            logger.error(f"Файл Excel поврежден или зашифрован {file_path.name}: {exc}")
            raise CorruptedExcelFileError(file_path=file_path.name, details=str(exc)) from exc
        except Exception as exc:
            logger.error(f"Ошибка при открытии Excel файла {file_path.name}: {exc}")
            raise DocumentConversionError(
                message=f"Не удалось открыть документ Excel '{file_path.name}': {exc}"
            ) from exc

        processed_sheets: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                sheet_md = self._process_single_sheet_with_docling(wb, sheet_name, file_path)
                if sheet_md:
                    processed_sheets.append(sheet_md)
        finally:
            wb.close()

        if not processed_sheets:
            raise EmptyDocumentError(file_path=file_path.name)

        return self.DELIMITER.join(processed_sheets)

    async def convert(self, file_path: Path) -> str:
        if not file_path.exists():
            raise DocumentFileNotFoundError(file_path=str(file_path))

        if not self.supports(file_path):
            raise UnsupportedFileFormatError(extension=file_path.suffix)

        start_time = time.perf_counter()
        logger.info(f"Старт конвертации Excel документа: {file_path.name}")

        try:
            markdown_content = await asyncio.to_thread(self._preprocess_excel, file_path)

            elapsed = time.perf_counter() - start_time
            logger.info(f"Excel файл {file_path.name} успешно конвертирован за {elapsed:.2f} c")
            return markdown_content

        except DocumentConversionError:
            raise
        except Exception as exc:
            logger.error(f"Непредвиденная ошибка конвертации Excel '{file_path.name}': {exc}")
            raise DocumentConversionError(
                message=f"Сбой при обработке Excel таблицы '{file_path.name}': {exc}"
            ) from exc