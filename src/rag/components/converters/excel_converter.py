import asyncio
from pathlib import Path
from typing import Set, Dict, Any, List
import openpyxl
import io

from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from docling.document_converter import DocumentConverter

from src.rag.components.converters import BaseDocumentConverter


import asyncio
import io
from pathlib import Path
from typing import Set

import openpyxl
from docling.datamodel.base_models import DocumentStream
from docling.document_converter import DocumentConverter


class ExcelConverter(BaseDocumentConverter):
    """
    Асинхронный конвертер Excel документов.
    Каждый лист изолируется, конвертируется через Docling с добавлением контекста,
    и объединяется через уникальный разделитель для сохранения единого типа возврата (str).
    """
    SUPPORTED_EXTENSIONS: Set[str] = {".xlsx", ".xls"}

    def __init__(self):
        self._docling_converter = DocumentConverter()

    def _process_single_sheet_with_docling(
        self, 
        wb_original: openpyxl.Workbook, 
        sheet_name: str, 
        file_path: Path
    ) -> str | None:
        """
        Изолирует один лист во временный файловый поток (BytesIO),
        прогоняет его через Docling и добавляет контекстный заголовок.
        """
        ws = wb_original[sheet_name]

        # Игнорируем скрытые и пустые листы
        if ws.sheet_state != 'visible':
            return None

        if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None:
            return None

        # Создаем чистую книгу Excel только с текущим листом
        single_sheet_wb = openpyxl.Workbook()
        single_sheet_wb.remove(single_sheet_wb.active)

        new_ws = single_sheet_wb.create_sheet(title=sheet_name)
        for row in ws.iter_rows(values_only=False):
            new_ws.append([cell.value for cell in row])

        buffer = io.BytesIO()
        single_sheet_wb.save(buffer)
        buffer.seek(0)
        file_name = file_path.name

        try:
            # Для передачи BytesIO в Docling используем DocumentStream
            doc_stream = DocumentStream(name=f"{sheet_name}.xlsx", stream=buffer)
            docling_result = self._docling_converter.convert(doc_stream)
            raw_markdown = docling_result.document.export_to_markdown()

            # Формируем контекстуальный заголовок (H1 и H2 для MarkdownHeaderTextSplitter)
            context_header = (
                f"# Файл: {file_name}\n"
                f"## Лист: {sheet_name}\n\n"
                f"> Context: Данные таблицы относятся к документу '{file_name}', раздел '{sheet_name}'.\n\n"
            )

            return context_header + raw_markdown
        finally:
            buffer.close()
            single_sheet_wb.close()

    def _preprocess_excel(self, file_path: Path) -> str:
        """
        Синхронный обход всех листов Excel и сборка в единую строку через DELIMITER.
        """
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        processed_sheets: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                sheet_md = self._process_single_sheet_with_docling(wb, sheet_name, file_path)
                if sheet_md:
                    processed_sheets.append(sheet_md)
        finally:
            wb.close()

        # Возвращаем ВСЕГДА единую строку (str)
        return self.DELIMITER.join(processed_sheets)

    async def convert(self, file_path: Path) -> str:
        """
        Гарантированно возвращает str в соответствии с интерфейсом BaseDocumentConverter.
        """
        loop = asyncio.get_running_loop()

        try:
            markdown_content = await loop.run_in_executor(
                None, 
                self._preprocess_excel, 
                file_path
            )

            return markdown_content

        except Exception as e:
            raise e